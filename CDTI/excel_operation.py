# -*- coding: utf-8 -*-
# author:xls
"""
    操作office对象类，主要用于读写execl、word，支撑报告自动生成功能
"""
import openpyxl
import csv

def get_csv_data(csv_filename):
    '''
    从csv文件中获取list[list[]]对象
    :param csv_filename: csv文件路径
    :return:list[list[]]
    '''
    csv_file = open(csv_filename, encoding='UTF-8')
    reader = csv.reader(csv_file)
    data = list(reader)
    return data



def read_data_from_excel(excel_filename, sheet_name):
    '''
    获取excel文本内容,返回一个list[list[]]对象
    :param excel_filename: excel文件路径
    :param sheet_name: sheet名称
    :return:
    '''
    result = []
    wb = openpyxl.load_workbook(excel_filename, data_only=True)
    if sheet_name == '':
        sheet = wb.get_active_sheet()
    else:
        sheet = wb.get_sheet_by_name(sheet_name)
    row_num = sheet.max_row
    col_num = sheet.max_column
    for row_index in range(row_num):
        row_list = []
        for col_index in range(col_num):
            row_list.append(sheet.cell(row=row_index+1, column=col_index+1).value)
        result.append(row_list)
    return result

def write_data_to_excel(excel_filename, sheet_name, location, data):
    '''
    向指定的excel文件写数据
    :param excel_filename: excel文件名
    :param sheet_name: sheet名称
    :param location: 表格位置 example 'A1' 'C3'
    :param data: 数据值
    :return:
    '''
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb.get_sheet_by_name(sheet_name)
    sheet[location] = data
    wb.save(excel_filename)



mylist = read_data_from_excel("UA-CDS接口设计V3.xlsx","命令映射关系")
map_dict = {}
set_list = []
for item in mylist:
    #set_list.append([item[1],item[3]])
    map_dict[item[0]]=item[2]
print(map_dict)

